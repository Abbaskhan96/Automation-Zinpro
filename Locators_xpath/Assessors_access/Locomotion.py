from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
import time
import openpyxl as o



class Locomotion_score():
   
    

    def __init__(self,driver, group):
       self.driver = driver
       #//input[@type='number' and @inputmode = 'numeric' and @class = 'score-count score-count--sm']
       #self.file = o.load_workbook("C:/Users/Muhammad Abbas Khan/Source/Repos/Automation-Zinpro/Excell_files/Locators.xlsx")
       self.file = o.load_workbook("C:/Users/CDC.CDC-PC/source/repos/Automation-Zinpro/Excell_Files/Locators.xlsx")
       self.sheet = self.file["Locomotion"]
       

       locom_dict = {
           "closed": self.closed_excell,
           "faroff": self.faroff_excell,
           "fresh": self.fresh_excell,
           "lactating": self.lactating_excell
           }
       locom_dict[group]()
       
       
       
    def closed_excell(self):
       score1 = str(self.sheet.cell(8, 6).value)
       score2 = str(self.sheet.cell(9, 6).value)
       score3 = str(self.sheet.cell(10, 6).value)
       score4 = str(self.sheet.cell(11, 6).value)
       score5 = str(self.sheet.cell(12, 6).value)
       
       self.driver.find_element(By.XPATH,"//*[@id='top']/main/div[1]/div[1]/div/div[2]/div[3]/div[6]/div[1]/div[2]/div[2]/input").send_keys(score1)
       self.driver.find_element(By.XPATH,f"//*[@id='top']/main/div[1]/div[1]/div/div[2]/div[3]/div[6]/div[2]/input").send_keys(score2) 
       self.driver.find_element(By.XPATH,f"//*[@id='top']/main/div[1]/div[1]/div/div[2]/div[3]/div[6]/div[3]/input").send_keys(score3)
       self.driver.find_element(By.XPATH,f"//*[@id='top']/main/div[1]/div[1]/div/div[2]/div[3]/div[6]/div[4]/input").send_keys(score4)
       self.driver.find_element(By.XPATH,f"//*[@id='top']/main/div[1]/div[1]/div/div[2]/div[3]/div[6]/div[5]/input").send_keys(score5)


    def faroff_excell(self):
       self.driver.find_element(By.XPATH,"//input[@value='confined_cattle']").click()
       score1 = str(self.sheet.cell(8, 8).value)
       score2 = str(self.sheet.cell(9, 8).value)
       score3 = str(self.sheet.cell(10, 8).value)
       score4 = str(self.sheet.cell(11, 8).value)
       score5 = str(self.sheet.cell(12, 8).value)
       
       self.driver.find_element(By.XPATH,"//*[@id='top']/main/div[1]/div[1]/div/div[2]/div[3]/div[6]/div[1]/div[2]/div[2]/input").send_keys(score1)
       self.driver.find_element(By.XPATH,f"//*[@id='top']/main/div[1]/div[1]/div/div[2]/div[3]/div[6]/div[2]/input").send_keys(score2) 
       self.driver.find_element(By.XPATH,f"//*[@id='top']/main/div[1]/div[1]/div/div[2]/div[3]/div[6]/div[3]/input").send_keys(score3)
       self.driver.find_element(By.XPATH,f"//*[@id='top']/main/div[1]/div[1]/div/div[2]/div[3]/div[6]/div[4]/input").send_keys(score4)
       self.driver.find_element(By.XPATH,f"//*[@id='top']/main/div[1]/div[1]/div/div[2]/div[3]/div[6]/div[5]/input").send_keys(score5)


    def fresh_excell(self):
       self.driver.find_element(By.XPATH,"//input[@value='confined_cattle']").click()
       self.driver.find_element(By.XPATH, "//label[@for='on']").click()
       score1 = str(self.sheet.cell(8, 10).value)
       score2 = str(self.sheet.cell(9, 10).value)
       score3 = str(self.sheet.cell(10, 10).value)
       score4 = str(self.sheet.cell(11, 10).value)
       score5 = str(self.sheet.cell(12, 10).value)
       
       self.driver.find_element(By.XPATH,"//*[@id='top']/main/div[1]/div[1]/div/div[2]/div[3]/div[6]/div[1]/div[2]/div[2]/input").send_keys(score1)
       self.driver.find_element(By.XPATH,f"//*[@id='top']/main/div[1]/div[1]/div/div[2]/div[3]/div[6]/div[2]/input").send_keys(score2) 
       self.driver.find_element(By.XPATH,f"//*[@id='top']/main/div[1]/div[1]/div/div[2]/div[3]/div[6]/div[3]/input").send_keys(score3)
       self.driver.find_element(By.XPATH,f"//*[@id='top']/main/div[1]/div[1]/div/div[2]/div[3]/div[6]/div[4]/input").send_keys(score4)
       self.driver.find_element(By.XPATH,f"//*[@id='top']/main/div[1]/div[1]/div/div[2]/div[3]/div[6]/div[5]/input").send_keys(score5)


    def lactating_excell(self):
       self.driver.find_element(By.XPATH,"//input[@value='grazing_cattle']").click()
       self.driver.find_element(By.XPATH, "//label[@for='on']").click()
       score1 = str(self.sheet.cell(8, 12).value)
       score2 = str(self.sheet.cell(9, 12).value)
       score3 = str(self.sheet.cell(10, 12).value)
       score4 = str(self.sheet.cell(11, 12).value)
       score5 = str(self.sheet.cell(12, 12).value)
       
       self.driver.find_element(By.XPATH,"//*[@id='top']/main/div[1]/div[1]/div/div[2]/div[3]/div[6]/div[1]/div[2]/div[2]/input").send_keys(score1)
       self.driver.find_element(By.XPATH,f"//*[@id='top']/main/div[1]/div[1]/div/div[2]/div[3]/div[6]/div[2]/input").send_keys(score2) 
       self.driver.find_element(By.XPATH,f"//*[@id='top']/main/div[1]/div[1]/div/div[2]/div[3]/div[6]/div[3]/input").send_keys(score3)
       self.driver.find_element(By.XPATH,f"//*[@id='top']/main/div[1]/div[1]/div/div[2]/div[3]/div[6]/div[4]/input").send_keys(score4)
       self.driver.find_element(By.XPATH,f"//*[@id='top']/main/div[1]/div[1]/div/div[2]/div[3]/div[6]/div[5]/input").send_keys(score5)

    