from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
import openpyxl
import time


class Dirt_alot():
    
    #|------------------HeadLock_FeedSpace Atleast_one---------------------
    headlock_FS_no="/html/body/div/main/div[1]/div[1]/div/div[2]/div[3]/div[2]/div[1]/label"
    headlock_FS_yes="/html/body/div/main/div[1]/div[1]/div/div[2]/div[3]/div[2]/div[2]/label"

    #|------------------Concreate_FeedSpace Scraped Atleast_one---------------------
    Scrapped_FS_no="/html/body/div/main/div[1]/div[1]/div/div[2]/div[4]/div[2]/div[1]/label"
    Scrapped_FS_yes="/html/body/div/main/div[1]/div[1]/div/div[2]/div[4]/div[2]/div[2]/label"
    
    #|------------------Bedding Dry Free of Wet---------------------
    bedding_Dry_no="/html/body/div/main/div[1]/div[1]/div/div[2]/div[5]/div[2]/div[1]/label"
    bedding_Dry_yes="/html/body/div/main/div[1]/div[1]/div/div[2]/div[5]/div[2]/div[2]/label"
    
    #|------------------atleast 7.5 cm loose dirt in bedding area`---------------------
    looseDirt_no="/html/body/div/main/div[1]/div[1]/div/div[2]/div[6]/div[2]/div[1]/label"
    looseDirt_yes="/html/body/div/main/div[1]/div[1]/div/div[2]/div[6]/div[2]/div[2]/label"

    #|------------------Knee Drop Test---------------------
    KneeDrop_no="/html/body/div/main/div[1]/div[1]/div/div[2]/div[7]/div[2]/div[1]/label"
    KneeDrop_yes="/html/body/div/main/div[1]/div[1]/div/div[2]/div[7]/div[2]/div[2]/label"

    
#    path_yes= {
#            headlock_FS_yes: "/html/body/div/main/div[1]/div[1]/div/div[2]/div[3]/div[2]/div[2]/label",
#            Scrapped_FS_yes : "/html/body/div/main/div[1]/div[1]/div/div[2]/div[4]/div[2]/div[2]/label",
#            bedding_Dry_yes : "/html/body/div/main/div[1]/div[1]/div/div[2]/div[5]/div[2]/div[2]/label",
#            looseDirt_yes : "/html/body/div/main/div[1]/div[1]/div/div[2]/div[6]/div[2]/div[2]/label",
#            KneeDrop_yes : "/html/body/div/main/div[1]/div[1]/div/div[2]/div[7]/div[2]/div[2]/label"
#            }
#
#    path_no = {
#            headlock_FS_no: "/html/body/div/main/div[1]/div[1]/div/div[2]/div[3]/div[2]/div[1]/label",
#            Scrapped_FS_no : "/html/body/div/main/div[1]/div[1]/div/div[2]/div[4]/div[2]/div[1]/label",
#            bedding_Dry_no : "/html/body/div/main/div[1]/div[1]/div/div[2]/div[5]/div[2]/div[1]/label",
#            looseDirt_no : "/html/body/div/main/div[1]/div[1]/div/div[2]/div[6]/div[2]/div[1]/label",
#            KneeDrop_no : "/html/body/div/main/div[1]/div[1]/div/div[2]/div[7]/div[2]/div[1]/label"
#        }
#
#    path_yes_no = {
#            headlock_FS_yes: "/html/body/div/main/div[1]/div[1]/div/div[2]/div[3]/div[2]/div[2]/label",
#            Scrapped_FS_yes : "/html/body/div/main/div[1]/div[1]/div/div[2]/div[4]/div[2]/div[2]/label",
#            bedding_Dry_yes : "/html/body/div/main/div[1]/div[1]/div/div[2]/div[5]/div[2]/div[2]/label",
#            looseDirt_no : "/html/body/div/main/div[1]/div[1]/div/div[2]/div[6]/div[2]/div[1]/label",
#            KneeDrop_no : "/html/body/div/main/div[1]/div[1]/div/div[2]/div[7]/div[2]/div[1]/label"
#        }
#
#    path_each_yes_no={
#            headlock_FS_yes: "/html/body/div/main/div[1]/div[1]/div/div[2]/div[3]/div[2]/div[2]/label",
#            Scrapped_FS_no : "/html/body/div/main/div[1]/div[1]/div/div[2]/div[4]/div[2]/div[1]/label",
#            bedding_Dry_yes : "/html/body/div/main/div[1]/div[1]/div/div[2]/div[5]/div[2]/div[2]/label",
#            looseDirt_no : "/html/body/div/main/div[1]/div[1]/div/div[2]/div[6]/div[2]/div[1]/label",
#            KneeDrop_yes : "/html/body/div/main/div[1]/div[1]/div/div[2]/div[7]/div[2]/div[2]/label"
#        }
    def __init__(self,driver, group):
        self.driver = driver
        # file= openpyxl.load_workbook("C:/Users/CDC.CDC-PC/source/repos/Automation-Zinpro/Excell_Files/Locators.xlsx")
        self.file = openpyxl.load_workbook("C:/Users/Muhammad Abbas Khan/Source/Repos/Automation-Zinpro/Excell_files/Locators.xlsx")
        self.sheet = self.file["DirtAlot"]
        Dirt_dict = {
           "closed": self.closed_Dirt,
           "faroff": self.faroff_Dirt,
           "fresh": self.fresh_Dirt,
           "lactating": self.lactating_Dirt
           }
        

        self._global = "Dirt_alot"
        self.data_click = {
                        "yes" : (Dirt_alot.headlock_FS_yes, Dirt_alot.Scrapped_FS_yes, Dirt_alot.bedding_Dry_yes, Dirt_alot.looseDirt_yes, Dirt_alot.KneeDrop_yes),
                        "no" : (Dirt_alot.headlock_FS_no, Dirt_alot.Scrapped_FS_no,Dirt_alot.bedding_Dry_no, Dirt_alot.looseDirt_no, Dirt_alot.KneeDrop_no)
                        }
        
        Dirt_dict[group]()




    def data_sheet(self, start_rows,last_rows, column):
        path_yes = None
        path_yes = []
        i,j = start_rows,0
        self.driver.find_element(By.XPATH, "//input[@type='number']").clear()
        self.driver.find_element(By.XPATH, "//input[@type='number']").send_keys(self.sheet.cell(start_rows-1, column).value)

        while (i != last_rows+1 and j!=5):
            """ store the data from conditions of excell self.sheet either the excell dropdown (yes/no) and store in path_yes"""
            print(i, j)
            print(self.sheet.cell(i,column).value)
            _get_click_data = self.data_click[self.sheet.cell(i,column).value][j]
            path_yes.append(_get_click_data)
            print(path_yes,"\n")
            i=i+1
            j=j+1
        return path_yes



    def closed_Dirt(self):
#        self.func = lambda : [i for i in self.path_yes]
#        [self.driver.find_element(By.XPATH, i).click() for i in self.func()]
#
#        self.corral_area = self.driver.find_element(By.XPATH,"//input[@type='number']")
#        self.corral_area.clear()
#        self.corral_area.send_keys("1500")
#        time.sleep(3)
        
         loc = self.data_sheet(7,12,7)
         for i in loc:
             # print(i,"\n")
             self.driver.find_element(By.XPATH, i).click()
         time.sleep(2)


    def faroff_Dirt(self):
#        self.func = lambda : [i for i in self.path_no]
#        [self.driver.find_element(By.XPATH, i).click() for i in self.func()]
#
#        self.corral_area = self.driver.find_element(By.XPATH,"//input[@type='number']")
#        self.corral_area.clear()
#        self.corral_area.send_keys("500")
#        time.sleep(3)
        loc = self.data_sheet(7,12,9)
        for i in loc:
             # print(i,"\n")
             self.driver.find_element(By.XPATH, i).click()
        time.sleep(2)

    def fresh_Dirt(self):
#        self.func = lambda : [i for i in self.path_yes_no]
#        [self.driver.find_element(By.XPATH, i).click() for i in self.func()]
#
#        self.corral_area = self.driver.find_element(By.XPATH,"//input[@type='number']")
#        self.corral_area.clear()
#        self.corral_area.send_keys("1000")
#        time.sleep(3)
        loc = self.data_sheet(7,12,11)
        for i in loc:
             # print(i,"\n")
             self.driver.find_element(By.XPATH, i).click()
        time.sleep(2)


    def lactating_Dirt(self):
#        self.func = lambda : [i for i in self.path_each_yes_no]
#        [self.driver.find_element(By.XPATH, i).click() for i in self.func()]
#
#        self.corral_area = self.driver.find_element(By.XPATH,"//input[@type='number']")
#        self.corral_area.clear()
#        self.corral_area.send_keys("300")
#        time.sleep(3)
        loc = self.data_sheet(7,12,9)
        for i in loc:
             # print(i,"\n")
             self.driver.find_element(By.XPATH, i).click()
        time.sleep(2)


