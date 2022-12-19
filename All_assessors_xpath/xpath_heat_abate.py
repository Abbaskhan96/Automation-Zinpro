from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
import openpyxl 
import time
class Assessors_HeatAbatement_xpath():
        accessible_water_pen = "//div[@class='col-lg-12']//div[@class='form-group mb-0 mb-sm-1']//input[@type='number']"
        
            


#===================================<<<======| Heat- Abatement |======>>>======================================        
        accessible_water_pen = "//div[@class='col-lg-12']//div[@class='form-group mb-0 mb-sm-1']//input[@type='number']"
 
 #__________________________________________| W A T E R |____________________________________________________
 #----------------------------------------Water_Availablity---------|
        WA_yes = "//div[4][@class='row']//child::div[2]//child::div[2]//child::label"
        WA_no = "//div[4][@class='row']//child::div[2]//child::div[1]//child::label"
        my_dict = {WA_yes : WA_yes,  
                    WA_no : WA_no} 
 #----- ---------Area arount the Water is free of muds/stones-----------|
        Warea_yes = "//div[5][@class='row']//child::div[2]//child::div[2]//child::label"
        Warea_no = "//div[5][@class='row']//child::div[2]//child::div[2]//child::label"
#``````` ``````````````````````````````````````````````````````````````````````````````````````````````````````
#______ ______________________________________| F A N S |_______________________________________________________
#------ --------------------------------Fans Located in Feed Area?--------|
        FL_FA_no = "//div[6][@class='row']//div[3]//child::div[2]//child::div[1]//label"  
        FL_FA_yes = "//div[6][@class='row']//div[3]//child::div[2]//child::div[2]//label"
        HVLS_FA_yes="//div[6][@class='row']//div[2]//child::div[2]//child::div[2]//label"
        HVLS_FA_no="//div[6][@class='row']//div[2]//child::div[2]//child::div[1]//label"
#------ ---------Fans'Temperature?-----------------------|
        Fan_FA_temp = "/html[1]/body[1]/div[1]/main[1]/div[1]/div[1]/div[1]/div[2]/div[7]/div[1]/div[1]/input[1]"
#------ ---------Fan's Diameter? (SELECT)-------------------------|
        Fan_FA_diamter = "//select[@name='feed_bunk_area_fan_diameter']"
#------ ---------Fan's spacing?--------------------------|
        Fan_FA_space = "/html[1]/body[1]/div[1]/main[1]/div[1]/div[1]/div[1]/div[2]/div[7]/div[3]/div[1]/input[1]"
#``````` `````````````````````````````````````````````````````````````````````````````````````````````````````
#______ ______________________________|R E S T I N G _ A R E A |________________________________________
#------ ---------------------------------Fan's Located in Resting---------------------------------------------- 
        FL_RA_no = "//*[@id='top']/main/div[1]/div/div/div[2]/div[8]/div[2]/div[1]/div[2]/div[1]/label"
        FL_RA_yes = "//*[@id='top']/main/div[1]/div/div/div[2]/div[8]/div[2]/div[1]/div[2]/div[2]/label"
        HVLS_RA_yes="//*[@id='top']/main/div[1]/div/div/div[2]/div[8]/div[2]/div[2]/div[2]/div[2]/label"
        HVLS_RA_no="//*[@id='top']/main/div[1]/div/div/div[2]/div[8]/div[2]/div[2]/div[2]/div[1]/label"
#------ ---------Fans'Temperature?-----------------------|
        Fan_RA_temp = "/html[1]/body[1]/div[1]/main[1]/div[1]/div[1]/div[1]/div[2]/div[9]/div[1]/div[1]/input[1]"
#------ ---------Fan's Diameter (SELECT)?-------------------------|
        Fan_RA_diamter = "//select[@name='resting_area_paddock_fan_diameter']"
#------ ---------Fan's spacing?--------------------------|
        Fan_RA_space = "/html[1]/body[1]/div[1]/main[1]/div[1]/div[1]/div[1]/div[2]/div[9]/div[3]/div[1]/input[1]"
        
#``````` `````````````````````````````````````````````````````````````````````````````````````````````````````
#______ ______________________________________| S O A K E R S |________________________________________
        
#------ ---------------------------------Soakers Located in Feeding Area---------------------------------------------- 
        Soak_FA_no = "//*[@id='top']/main/div[1]/div/div/div[2]/div[10]/div[2]/div[1]/div[2]/div[1]/label"
        Soak_FA_yes = "//*[@id='top']/main/div[1]/div/div/div[2]/div[10]/div[2]/div[1]/div[2]/div[2]/label"
#------ ---------------------------------Soakers Located in Resting Area-----------------------------
        Soak_RA_no = "//*[@id='top']/main/div[1]/div[1]/div/div[2]/div[10]/div[2]/div[2]/div[2]/div[1]/label"
        Soak_RA_yes = "//*[@id='top']/main/div[1]/div[1]/div/div[2]/div[10]/div[2]/div[2]/div[2]/div[2]/label"
#------ ---------------------------------water droplet effecting cows------------------------------
        drop_no= "//*[@id='top']/main/div[1]/div[1]/div/div[2]/div[10]/div[2]/div[3]/div[2]/div[1]/label"
        drop_yes="//*[@id='top']/main/div[1]/div[1]/div/div[2]/div[10]/div[2]/div[3]/div[2]/div[2]/label"
#------ ---------Soak Temperature?----------------------|
        Soak_temp = "/html[1]/body[1]/div[1]/main[1]/div[1]/div[1]/div[1]/div[2]/div[11]/div[1]/div[1]/input[1]"
#------ ---------Soak Duration?-------------------------|
        Soak_time = "/html[1]/body[1]/div[1]/main[1]/div[1]/div[1]/div[1]/div[2]/div[11]/div[2]/div[1]/input[1]"
#------ ---------Soak Frequency?------------------------|
        Soak_freq = "/html[1]/body[1]/div[1]/main[1]/div[1]/div[1]/div[1]/div[2]/div[11]/div[3]/div[1]/input[1]"
#------ ---------------------------------Frequency of soaking vary with Temperature------------------------------
        SoakFreq_VT_no= "//*[@id='top']/main/div[1]/div[1]/div/div[2]/div[11]/div[4]/div/div[2]/div[1]/label"
        SoakFreq_VT_yes= "//*[@id='top']/main/div[1]/div[1]/div/div[2]/div[11]/div[4]/div/div[2]/div[2]/label"
#``````` ``````````````````````````````````````````````````````````````````````````````````````````````````````````
##_____ ______________________________________| S H A D E |________________________________________
        
#------ ---------------------------------Soakers Located in Feeding Area---------------------------------------------- 
        Shade_FA_no = "//*[@id='top']/main/div[1]/div[1]/div/div[2]/div[12]/div[2]/div[1]/div[2]/div[1]/label"
        Shade_FA_yes = "//*[@id='top']/main/div[1]/div[1]/div/div[2]/div[12]/div[2]/div[1]/div[2]/div[2]/label"
#------ ---------------------------------Soakers Located in Resting Area---------------------------------------------- 
        Shade_RA_no = "//*[@id='top']/main/div[1]/div[1]/div/div[2]/div[12]/div[2]/div[2]/div[2]/div[1]/label"
        Shade_RA_yes = "//*[@id='top']/main/div[1]/div[1]/div/div[2]/div[12]/div[2]/div[2]/div[2]/div[2]/label"
#------ ---------------------------------AirCooling units used in Resting Area------------------------------------
        Shade_unit_no="//*[@id='top']/main/div[1]/div[1]/div/div[2]/div[12]/div[2]/div[3]/div[2]/div[1]/label"
        Shada_unit_yes="//*[@id='top']/main/div[1]/div[1]/div/div[2]/div[12]/div[2]/div[3]/div[2]/div[2]/label"
#====== ================================================================================================================================================
        
        
        
        select_fields = [
            Fan_FA_diamter,
            Fan_RA_diamter
            ]
#       
#        path_yes = {
#             WA_sheet.cell(10,7).value : WA_sheet.cell(10,7).value, 
#            Warea_sheet.cell(11,7).value : Warea_sheet.cell(11,7).value,
#                 FL_FA_sheet.cell(15,7).value : FL_FA_sheet.cell(15,7),
#                 HVLS_FA_sheet.cell(16,7).value : HVLS_FA_sheet.cell(16,7).value,
#                 FL_RA_sheet.cell(22,7).value : FL_RA_sheet.cell(22,7).value,
#                 HVLS_RA_sheet.cell(23,7).value : HVLS_RA_sheet.cell(23,7).value,
#                 Soak_FA_sheet.cell(29,7).value : Soak_FA_sheet.cell(29,7).value,
#                 Soak_RA_sheet.cell(30,7).value : Soak_RA_sheet.cell(30,7).value,
#                 drop_sheet.cell(31,7).value : drop_sheet.cell(31,7).value,
#                 SoakFreq_VT_sheet.cell(34,7).value : SoakFreq_VT_sheet.cell(34,7).value,
#                 Shade_FA_sheet.cell(38,7).value : Shade_FA_sheet.cell(38,7).value,
#                 Shade_RA_sheet.cell(39,7).value : Shade_RA_sheet.cell(39,7).value,
#                 Shada_unit_sheet.cell(40,7).value : Shada_unit_sheet.cell(40,7).value
#                }
#path_yes = {
#       'yes' : "yes-path",
#       'no' : 'no_path'
#    }
#
#path_yes = {
#    self.sheet.cell(10,7).value : {'yes':'yes_path', 'no':'no_path'}
#    }
#path_yes[self.sheet[10,7].value][self.sheet[10,7].value]
#[j[i] for i,j in path_yes.items]
#my_final_dict = {
#    WA_yes
#    }
        def __init__(self, driver):
            self.driver= driver
           # file= openpyxl.load_workbook("C:/Users/CDC.CDC-PC/source/repos/Automation-Zinpro/Excell_Files/Locators.xlsx")
            self.file = openpyxl.load_workbook("C:/Users/Muhammad Abbas Khan/Source/Repos/Automation-Zinpro/Excell_files/Locators.xlsx")
            self.sheet = self.file["HeatAbatement"]
            
            _global = Assessors_HeatAbatement_xpath
         #   self.driver.find_element(By.XPATH, _global.accessible_water_pen).send_keys(sheet.cell(9,7).value)
            
            self.data_click = { 'yes' : ( _global.WA_yes, _global.Warea_yes,
                                   _global.FL_FA_yes, _global.HVLS_FA_yes, 
                                   _global.FL_RA_yes ,_global.HVLS_RA_yes, 
                                   _global.Soak_FA_yes, 
                                   _global.Soak_RA_yes, _global.drop_yes,
                                   _global.Shade_FA_yes,
                                   _global.Shade_RA_yes,_global.Shada_unit_yes),
                           
                            'no' : (_global.WA_no, _global.Warea_no, 
                                    _global.FL_FA_no, _global.HVLS_FA_no, 
                                    _global.FL_RA_no, _global.HVLS_RA_no, 
                                    _global.Soak_FA_no, 
                                    _global.Soak_RA_no, _global.drop_no,
                                    _global.Shade_FA_no,
                                    _global.Shade_RA_no, _global.Shade_unit_no)}
            
            
            self.data_keys = {_global.FL_FA_yes : (_global.Fan_FA_temp, _global.Fan_FA_space),
                         _global.FL_RA_yes : (_global.Fan_RA_temp, _global.Fan_RA_space),
                         _global.Soak_RA_yes : (_global.Soak_temp, _global.Soak_freq)}

            self.data_select = {
                           _global.FL_FA_yes : _global.Fan_FA_diamter,
                           _global.FL_RA_yes : _global.Fan_RA_diamter
                            }
            
            """
            self.data_select = {_global.Fan_FA_diamter : 
                        }
           """ 
  
        def data_sheet(self,start_rows, last_rows,  column):    
            
            path_yes = None
            path_yes=[]
            i,j = start_rows,0
            
            _global = Assessors_HeatAbatement_xpath
            self.driver.find_element(By.XPATH,_global.accessible_water_pen).clear()
            self.driver.find_element(By.XPATH,_global.accessible_water_pen).send_keys(self.sheet.cell(start_rows - 1,column).value)
            
            while (i!=last_rows+1 and j!=12):
                
                
                

                """ store the data from conditions of excell self.sheet either the excell dropdown (yes/no) and store in path_yes"""
                _get_click_data = self.data_click[self.sheet.cell(i,column).value][j]

                print(f"{i},{j}","\n",_get_click_data,"\n","\n")
                path_yes.append(_get_click_data)
                
                """check for multi-dependent questions that after clicking fields temperature and time text field open"""
                if ((_get_click_data == _global.FL_FA_yes) or
                   (_get_click_data == _global.FL_RA_yes)  or
                   (_get_click_data == _global.Soak_RA_yes)):
                    
                    """ clicking on those elements first we are not waiting for end return of path_yes 
                     from this approach we will be able to send keys in Temp and time,space text fields"""
                    self.driver.find_element(By.XPATH, _get_click_data).click()
                    
                    """ get the list of Temperature and space"""
                    temp_space = self.data_keys[_get_click_data]
                   # print("send arrays")
                   # print(temp_space)
                    for k,l in zip(temp_space, range(i+11,i+13)):
                        
                      #  print("values")
                      #  print(self.sheet.cell(l,7).value)
                      # send_keys from Excell Cell
                     # sending keys from excell cell to Temp,Space text site field 
                        
                        self.driver.find_element(By.XPATH, k).send_keys(self.sheet.cell(l,column).value)
                    if _global.Soak_RA_yes in path_yes:
                        self.driver.find_element(By.XPATH, _global.Soak_time).send_keys(self.sheet.cell(30,column).value) 
                        self.driver.find_element(By.XPATH, _global.SoakFreq_VT_yes).click() 
                
                    """
                Checking that the Fans in Feed Area/Resting areas are available
                """
                    if ((_get_click_data == _global.FL_FA_yes) or
                    (_get_click_data == _global.FL_RA_yes)):
                        print(f"{i+20},{column}")
                        self.selecting_dropdown(self.data_select[_get_click_data], (self.sheet.cell(i+20,column).value), self.driver)
                        
              
                """ for skipping the dependent field if their parents answer is no"""
                if ((_get_click_data == _global.FL_FA_no) or 
                    (_get_click_data == _global.FL_RA_no) or 
                    (_get_click_data == _global.Soak_RA_no) or
                    (_get_click_data == _global.Shade_RA_no)):
                    i+=1 
                    j+=1
  
                i=i+1
                j=j+1

        
          
               
            print("Final list <Return>","\n")
            print(i for i in path_yes)
            
            """ returning the final list in Heat Abatment file """
            return path_yes
            



        def selecting_dropdown(self,xpath, drop_sheet_value, driver):
            #for i,j in zip(loc.select_fields,index):
            print("Excell Value","\n")
            print(drop_sheet_value)
            drop_field = Select(driver.find_element(By.XPATH,xpath))
            time.sleep(1)
            drop_field.select_by_visible_text(str(drop_sheet_value))
            time.sleep(2)
            return
#            for i,j in zip(range(10,16), range(6)):
#                #print(sheet.cell(i,7).value, data[sheet.cell(i,7).value][j])
#                #path_yes = data[sheet.cell(i,7).value][j]
#                if not data[0 or 1][FL_FA_yes or FL_RA_yes]
#                path_yes.append(data[sheet.cell(i,7).value][j])
#                print("path_yes: ", path_yes)
#            return path_yes
 #                     
 #       
 #       path_yes = {
 #                WA_yes : WA_yes,
 #                Warea_yes : Warea_yes,
 #                FL_FA_yes : FL_FA_yes,
 #                HVLS_FA_no : HVLS_FA_no,
 #                FL_RA_yes : FL_RA_yes,
 #                HVLS_RA_no : HVLS_RA_no,
 #                Soak_FA_yes : Soak_FA_yes,
 #                Soak_RA_yes : Soak_RA_yes,
 #                drop_yes : drop_yes,
 #                SoakFreq_VT_yes : SoakFreq_VT_yes,
 #                Shade_FA_yes : Shade_FA_yes,
 #                Shade_RA_yes : Shade_RA_yes,
 #                Shada_unit_yes : Shada_unit_yes
 #               }

            path_no = {
                 WA_no : WA_no,
                 Warea_no : Warea_no,
                 FL_FA_no : FL_FA_no,
                 FL_RA_no : FL_RA_no,
                 Soak_FA_no : Soak_FA_no,
                 Soak_RA_no : Soak_RA_no,
                 Shade_FA_no :Shade_FA_no,
                 Shade_RA_no : Shade_RA_no
                }
    
            path_N_A = {
                 WA_no : WA_no,
                 Warea_no : Warea_no,
                 FL_FA_no : FL_FA_no,
                 FL_RA_no : FL_RA_no,
                 Soak_FA_no : Soak_FA_no,
                 Soak_RA_no : Soak_RA_no,
                 Shade_FA_no : Shade_FA_no,
                 Shade_RA_no : Shade_RA_no,   
                }
    
            path_yes_no = {
                WA_no : WA_no,  
                Warea_no : Warea_no,
                FL_FA_yes : FL_FA_yes,
                HVLS_FA_yes: HVLS_FA_yes,
                FL_RA_yes : FL_RA_yes,
                HVLS_RA_yes : HVLS_RA_yes,
                Soak_FA_yes : Soak_FA_yes,
                Soak_RA_yes: Soak_RA_yes,
                drop_no: drop_no,
                SoakFreq_VT_no : SoakFreq_VT_no,
                Shade_FA_no : Shade_FA_no,
                Shade_RA_yes : Shade_RA_yes,
                Shade_unit_no : Shade_unit_no
                }